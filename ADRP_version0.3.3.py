import openpyxl
import os
import shutil
import time
import tkinter as tk
from tkinter import messagebox, ttk
from tkinter.filedialog import askopenfile
from tkinter.ttk import *
import zipfile as zf

mainwdw = tk.Tk()
mainwdw.title("Automated Drawing Retrieval Program")
s_wdw = ttk.Style()
s_wdw.theme_use('xpnative')
fontsize = ("Arial", 12)

newdwglist = []
qwert = 0

lbl1 = tk.Label(mainwdw, text="Drawing list", font=fontsize)
lbl2 = tk.Label(mainwdw, text="Output location", font=fontsize)
sourcelink = tk.Entry(mainwdw, bd=5, width=100)
outputlink = tk.Entry(mainwdw, bd=5, width=100)
lbl1a_text = tk.StringVar()
lbl1a_text.set("BOM is not loaded yet.")
lbl1a = tk.Label(mainwdw, textvariable=lbl1a_text, font=fontsize)
lbl3 = tk.Label(mainwdw, text="Source location", font=fontsize)

def browseforbom():
    global newdwglist, qwert, lbl1a_text
    lbl1a_text.set("Loading BOM...")
    bomfile = askopenfile(parent=mainwdw, mode='rb', title='Choose your BOM', filetypes=[("Excel file", "*.xlsx")])
    if bomfile:
        wb = openpyxl.load_workbook(bomfile, data_only=True)
        sheet = wb.active
        m_row = sheet.max_row
        drawing_list = []
        rev_list = []
        for i in range(1, m_row + 1):
            dwgnum = sheet.cell(row=i, column=2)
            drawing_list.append(dwgnum.value)
            revnum = sheet.cell(row=i, column=3)
            rev_list.append(revnum.value)

        for i in range(len(drawing_list)):
            a = str(drawing_list[i])
            b = str(rev_list[i])
            if a != "Nonerev_None.dwg":
                newdwglist.append(a + '_rev' + b + '.dwg')

        qwert = len(newdwglist)
    if len(newdwglist) > 1:
        lbl1a_text.set(f"BOM is loaded. There are {qwert} drawings found.")
    else:
        lbl1a_text.set("BOM is still not loaded")


def transfer():
    stime = time.perf_counter()
    source = sourcelink.get()
    destlink = outputlink.get()
    missingdwgs = []
    for i, dwgfile in enumerate(newdwglist, start=0):
        dwgfile = newdwglist[i]
        for roots, direcs, files in os.walk(source):
            for dwg in files:
                if dwg == dwgfile:
                    shutil.copy(dwgfile, destlink)
                else:
                    missingdwgs.append(dwg)
    etime = time.perf_counter()
    ttime = etime - stime
    messagebox.showinfo("Time taken is...", f"Your transfer took {round(ttime, 1)} seconds")
    if len(missingdwgs) != 0:
        messagebox.showinfo(f"Drawings not found: {missingdwgs}")

def validatelink():
    getlink_1 = sourcelink.get()
    getlink_2 = outputlink.get()
    if os.path.exists(getlink_1) and os.path.exists(getlink_2):
        transferbtn['state'] = 'active'
        messagebox.showinfo("Result", "Your link is valid!")
    else:
        messagebox.showinfo("Result", "Please try again.")

def closewdw():
    mainwdw.destroy()

validatebtn = tk.Button(mainwdw, text="Validate", command=validatelink, font=fontsize, width=20)

transferbtn = tk.Button(mainwdw, text="Transfer", command=transfer, font=fontsize, width=50)
transferbtn['state'] = 'disable'

brbtn = tk.Button(mainwdw, text="Browse", command=browseforbom, font=fontsize, width= 15)

lbl1.grid(row=1, column=1, pady=10, padx=10)
lbl1a.grid(row=1,column=2, pady=10, padx=10)
brbtn.grid(row=1, column=3, sticky='we', padx=25)

lbl3.grid(row=2, column=1, pady=10, padx=10)
sourcelink.grid(row=2,column=2, pady=10, padx=10)

lbl2.grid(row=3, column=1, pady=10, padx=10)
outputlink.grid(row=3, column=2, pady=10, padx=10)

validatebtn.grid(row=4, column=1, pady=1, padx=10)
transferbtn.grid(row=4, column=2, pady=10, padx=10)

extbtn_m = tk.Button(mainwdw, text="Exit", command=closewdw, font=fontsize)
extbtn_m.grid(row=10, column=2, pady=5)

mainwdw.bind('<Escape>', lambda e: closewdw())

if __name__ == '__main__':
    mainwdw.mainloop()
