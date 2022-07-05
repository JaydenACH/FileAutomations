import os
import openpyxl
import collections
from pathlib import Path
import tkinter as tk
import tkinter.messagebox as msgbox
from datetime import date
import subprocess

# The function of this code is to loop through the folder, separate out the drawing numbers & revision numbers,
# and check for drawings with multiple revisions, keep the largest revision of that drawing numbers and
# write the values into a new Excel file.
# _____________________________________________________________________________________________________________

class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()

        # Window Title Bar
        self.title("Get Drawing Revisions...")
        # Window positioning
        self.height = 300
        self.width = 800
        self.screenwidth = self.winfo_screenwidth()
        self.screenheight = self.winfo_screenheight()
        self.paddingw = (self.screenwidth/2) - (self.width/2)
        self.paddingh = (self.screenheight/2) - (self.height/2)
        self.geometry('%dx%d+%d+%d' % (self.width, self.height, self.paddingw, self.paddingh))

        # Window GUI Design
        self.fontsize = ('Arial Bold', 15)
        bgcolor = "#282C34"
        self.configure(background=bgcolor)
        self.copyright = u"\u00A9" "CopyRight by Jayden Ang"

        # Put in the directory that going to be the folder of all files
        self.drawingdir = os.getcwd()
        self.folderpath = "explorer "
        # Define some empty lists
        self.drawingcol = []
        self.revisioncol = []
        self.drawinglist = []

        self.input = date.today()
        self.filedate = self.input.strftime("%Y%m%d")

        self.label1 = tk.Label(self,
                               text="You are going to get the drawing revisions from",
                               background=bgcolor,
                               foreground="#FFFFFF",
                               font=self.fontsize)
        self.label2 = tk.Label(self,
                               text=self.drawingdir,
                               font=('Arial', 15, 'underline'),
                               foreground="#0000EE",
                               cursor="hand2",
                               background=bgcolor)

        self.button1 = tk.Button(self,
                                 text="Export Drawing Revisions",
                                 command=self.getrevision,
                                 height=500,
                                 cursor="hand2",
                                 bg=bgcolor,
                                 activebackground="#B5F9F2",
                                 foreground="#FFFFFF",
                                 font=self.fontsize,
                                 borderwidth=5)

        self.label3 = tk.Label(self,
                               text=self.copyright,
                               background=bgcolor,
                               foreground="#d9d9d9")

        self.label1.pack(side=tk.TOP, padx=50, pady=30)
        self.label2.pack(side=tk.TOP, padx=50)
        self.label2.bind("<Button>", lambda e: self.callback(self.drawingdir))
        self.label3.pack(side=tk.BOTTOM)
        self.button1.pack(side=tk.BOTTOM, padx=50, pady=20)

        self.bind('<Escape>', lambda e: exit())

    def getrevision(self):
        # Change directory from directory of this python file to the directory above
        os.chdir(self.drawingdir)

        # List out all files in folder and sub-folders
        # Files with extension of .dwg or .DWG will be added into the drawinglist list
        for roots, dirs, files in os.walk(self.drawingdir):
            for file in files:
                if file.endswith(".dwg") or file.endswith(".DWG"):
                    if file not in self.drawinglist:
                        self.drawinglist.append(file)

        # Sort the drawing list in descending order (A-Z) and get the length of list
        self.drawinglist.sort(reverse=False)
        length = len(self.drawinglist)

        # Loop through the whole list.
        # Drop the "_rev#.dwg" and add into drawingcol list.
        # Add # from rev into revisioncol list.
        for i in range(length):
            self.drawingcol.append(self.drawinglist[i][:-9])
            self.revisioncol.append(int(self.drawinglist[i][-5]))

        # Count how many revisions # for each drawing
        mylist = collections.Counter(self.drawingcol)

        # For each drawing (i) & number of drawings (j), if drawing is more than 1, check for larger revision number.
        # Keep the larger revision number & drop the smaller revision number from list of drawing & revision.
        # Length of list will -1 to prevent list looping out of index.
        for i, j in mylist.items():
            if j > 2:
                d = 1
                while d < length:
                    if self.drawingcol[d] == self.drawingcol[d - 1]:
                        if self.revisioncol[d] > self.revisioncol[d-1] or self.revisioncol[d] == self.revisioncol[d-1]:
                            del self.drawingcol[d-1]
                            del self.revisioncol[d-1]
                            del self.drawinglist[d-1]
                            length -= 1
                    d += 1

        # Create an Excel Workbook and activate the sheet, titled "Offline drawings revision"
        # Save the new created Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Offline drawings revision"
        wb.save(f'{self.filedate}_DrawingRevision.xlsx')

        # Loop through new drawingcol and write drawingcol into column A and revisioncol into column B.
        for i in range(len(self.drawingcol)):
            sheet.cell(row=i+1, column=1).value = self.drawingcol[i]
            sheet.cell(row=i+1, column=2).value = self.revisioncol[i]

        # Save workbook after writing everything.
        wb.save(f'{self.filedate}_DrawingRevision.xlsx')

        # Open up that excel after saving.
        if msgbox.askyesno("It's done", "Do you want to open the file?"):
            excelpath = Path(f'{self.filedate}_DrawingRevision.xlsx').resolve()
            os.system(f'start excel.exe "{excelpath}"')
            self.after(3000, self.exit())
        else:
            msgbox.showinfo("Message", "It's done anyway, go check it out.")
            self.after(3000, self.exit())

    def callback(self, link):
        link = self.folderpath + link
        subprocess.Popen(link)

    def exit(self):
        self.destroy()

if __name__ == '__main__':
    main = MainWindow()
    main.mainloop()
