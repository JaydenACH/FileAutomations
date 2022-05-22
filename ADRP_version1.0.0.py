import shutil
import tkinter as tk
import os
from tkinter.filedialog import askopenfile
import tkinter.messagebox as msgbox
import openpyxl
import time


class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.fontsize = ("Arial", 12)
        self.title("Automated Drawing Retrieval Program")
        self.label1 = tk.Label(self, text="Drawing list", font=self.fontsize)
        self.label2 = tk.Label(self, text="Source Location", font=self.fontsize)
        self.label3 = tk.Label(self, text="Output location", font=self.fontsize)

        self.sourcelink = tk.Entry(width=100, bd=5)
        self.destlink = tk.Entry(width=100, bd=5)

        self.varlabel_text = tk.StringVar()
        self.varlabel_text.set("No drawings found yet")
        self.varlabel = tk.Label(self, textvariable=self.varlabel_text, font=self.fontsize)

        validatebutton = tk.Button(self, text="Validate", command=self.validatelink, font=self.fontsize, width=20)
        self.transferbutton = tk.Button(self, text="Transfer", command=self.transfer, font=self.fontsize, width=20)
        self.transferbutton['state'] = 'disable'
        browsebutton = tk.Button(self, text="Browse", command=self.browse, font=self.fontsize, width=20)
        exitbutton = tk.Button(self, text="Exit", command=self.exit, font=self.fontsize, width=20)

        self.bind('<Escape>', lambda e: exit())

        self.label1.grid(row=1, column=1, pady=10, padx=10)
        self.varlabel.grid(row=1, column=2, pady=10, padx=10)
        browsebutton.grid(row=1, column=3, pady=10, padx=10)

        self.label2.grid(row=2, column=1, pady=10, padx=10)
        self.sourcelink.grid(row=2, column=2, pady=10, padx=10)

        self.label3.grid(row=3, column=1, pady=10, padx=10)
        self.destlink.grid(row=3, column=2, pady=10, padx=10)

        validatebutton.grid(row=4, column=1, pady=10, padx=10)
        self.transferbutton.grid(row=4, column=2, pady=10, padx=10)
        exitbutton.grid(row=5, column=2, pady=10, padx=10)

        self.newlistofdrawings = []

    def validatelink(self):
        getlink_1 = self.sourcelink.get()
        getlink_2 = self.destlink.get()
        if getlink_1 == "" or getlink_2 == "":
            msgbox.showwarning("Warning", "There is no link yet.")
        else:
            if not os.path.exists(getlink_1):
                msgbox.showerror("Error", f"This link is not found: {getlink_1}")
            elif not os.path.exists(getlink_2):
                msgbox.showerror("Error", f"This link is not found: {getlink_2}")
            else:
                msgbox.showinfo("Result", "Your link is valid!")
                self.transferbutton['state'] = 'active'

    def transfer(self):
        timeofstart = time.perf_counter()
        source = self.sourcelink.get()
        destination = self.destlink.get()
        missingdrawings = []
        for i, drawing in enumerate(self.newlistofdrawings, start=0):
            drawing = self.newlistofdrawings[i]
            for rts, direcs, files in os.walk(source):
                for dwg in files:
                    if dwg.lower() == drawing.lower():
                        os.chdir(rts)
                        shutil.copy(dwg, destination)

            for rts, direcs, files in os.walk(destination):
                if drawing not in files:
                    missingdrawings.append(drawing)

        timeofend = time.perf_counter()
        runtime = timeofend - timeofstart
        print(len(missingdrawings))
        print(missingdrawings)
        if len(missingdrawings) != 0:
            msgbox.showinfo("Missing drawings", f"Drawings not found: {missingdrawings}")
        else:
            msgbox.showinfo("Time taken is...", f"Your transfer took {round(runtime, 5)} seconds.")

    def browse(self):
        self.varlabel_text.set("Loading list of drawings")
        drawinglist = askopenfile(parent=self, title='Choose the file', mode='rb', filetypes=[("Excel file", "*.xlsx")])
        if drawinglist:
            workbook = openpyxl.load_workbook(drawinglist, data_only=True)
            sheet = workbook.active
            maxrow = sheet.max_row
            listofdrawings = []
            listofrevisions = []
            for i in range(1, maxrow + 1):
                drawingnumber = sheet.cell(row=i, column=1)
                listofdrawings.append(drawingnumber.value)
                revision = sheet.cell(row=i, column=2)
                listofrevisions.append(revision.value)

            for i in range(len(listofdrawings)):
                a = str(listofdrawings[i])
                b = str(listofrevisions[i])
                c = a + '_rev' + b + '.DWG'
                if c != "Nonerev_None.dwg":
                    self.newlistofdrawings.append(c)

        self.newlistofdrawings = list(set(self.newlistofdrawings))
        for i in self.newlistofdrawings:
            print(i)

        if len(self.newlistofdrawings) > 1:
            self.varlabel_text.set(f"There are {len(self.newlistofdrawings)} drawings found.")
        else:
            self.varlabel_text.set("Still no drawings found.")

    def exit(self):
        self.destroy()


if __name__ == '__main__':
    main = MainWindow()
    main.mainloop()
